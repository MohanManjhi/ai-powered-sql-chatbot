#!/usr/bin/env python3
"""Convert a pytest JUnit XML (`test_reports/junit_report.xml`) into an Excel report.

Writes `test_reports/unit_test_report.xlsx` with two sheets:
 - Summary: totals (tests, errors, failures, skipped, time, timestamp)
 - Details: one row per testcase with classname, name, time, status, message/trace

This script does not modify your application code.
"""
import os
import sys
from xml.etree import ElementTree as ET
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
except Exception as e:
    print("openpyxl is required. Activate your venv and install it: python -m pip install openpyxl")
    raise


ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", ".."))
XML_PATH = os.path.join(ROOT, "test_reports", "junit_report.xml")
OUT_XLSX = os.path.join(ROOT, "test_reports", "unit_test_report.xlsx")


def parse_junit(xml_path):
    if not os.path.exists(xml_path):
        raise FileNotFoundError(xml_path)
    tree = ET.parse(xml_path)
    root = tree.getroot()

    # Support both <testsuites> and <testsuite> root shapes
    suites = []
    if root.tag == "testsuites":
        suites = list(root.findall("testsuite"))
    elif root.tag == "testsuite":
        suites = [root]
    else:
        # fallback: find testsuite elements
        suites = root.findall(".//testsuite")

    summary = {
        "tests": 0,
        "errors": 0,
        "failures": 0,
        "skipped": 0,
        "time": 0.0,
        "timestamp": None,
    }

    details = []
    for suite in suites:
        summary["tests"] += int(suite.attrib.get("tests", 0))
        summary["errors"] += int(suite.attrib.get("errors", 0))
        summary["failures"] += int(suite.attrib.get("failures", 0))
        summary["skipped"] += int(suite.attrib.get("skipped", 0))
        try:
            summary["time"] += float(suite.attrib.get("time", 0.0))
        except Exception:
            pass
        if summary["timestamp"] is None:
            summary["timestamp"] = suite.attrib.get("timestamp")

        for case in suite.findall("testcase"):
            classname = case.attrib.get("classname", "")
            name = case.attrib.get("name", "")
            time = case.attrib.get("time", "")
            status = "passed"
            message = ""
            # check for failure/error/skip child
            for child in list(case):
                tag = child.tag.lower()
                if tag in ("failure", "error"):
                    status = tag
                    message = (child.attrib.get("message", "") + "\n" + (child.text or "")).strip()
                    break
                if tag == "skipped":
                    status = "skipped"
                    message = (child.attrib.get("message", "") + "\n" + (child.text or "")).strip()
                    break

            details.append({
                "classname": classname,
                "name": name,
                "time": time,
                "status": status,
                "message": message,
            })

    return summary, details


def write_excel(summary, details, out_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"

    ws.append(["Generated on", datetime.utcnow().isoformat() + "Z"]) 
    ws.append([])
    ws.append(["Total Tests", summary.get("tests")])
    ws.append(["Errors", summary.get("errors")])
    ws.append(["Failures", summary.get("failures")])
    ws.append(["Skipped", summary.get("skipped")])
    ws.append(["Total Time (s)", summary.get("time")])
    ws.append(["Timestamp", summary.get("timestamp")])

    # Details sheet
    ws2 = wb.create_sheet(title="Details")
    headers = ["Classname", "Test Name", "Time(s)", "Status", "Message/Trace"]
    ws2.append(headers)
    for d in details:
        ws2.append([d["classname"], d["name"], d["time"], d["status"], d["message"]])

    # Auto-column width (simple)
    for sheet in (ws, ws2):
        for col in sheet.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    val = str(cell.value) if cell.value is not None else ""
                except Exception:
                    val = ""
                if len(val) > max_length:
                    max_length = len(val)
            adjusted_width = min(max_length + 2, 100)
            sheet.column_dimensions[col_letter].width = adjusted_width

    wb.save(out_path)


def main():
    try:
        summary, details = parse_junit(XML_PATH)
    except FileNotFoundError:
        print(f"JUnit XML not found at {XML_PATH}")
        sys.exit(2)

    os.makedirs(os.path.dirname(OUT_XLSX), exist_ok=True)
    write_excel(summary, details, OUT_XLSX)
    print(f"Excel report written to: {OUT_XLSX}")


if __name__ == "__main__":
    main()
